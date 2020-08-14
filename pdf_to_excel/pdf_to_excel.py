# -*- coding: utf-8 -*-
"""
Created on Fri Jul 10 21:23:52 2020

@author: noelp
"""
from tabula import read_pdf
import pandas as pd
from openpyxl import Workbook
from openpyxl import worksheet
#%%

infile = 'catalogo2020.pdf'
output = 'catalogo2020.xlsx'

df = read_pdf(infile, encoding='utf-8', pages='all', multiple_tables=True)
#%%
wb = Workbook()

column = ['A','B','C','D','F','G','H']
dest_filename = output
for s in range(len(df)):
    ws = wb.create_sheet('hoja{}'.format(s))
    ##primera linea
    names = list(df[s].head())

    for r in range(len(names)):
        ws['{}1'.format(column[r])] = names[r]
    
    ###resto de la tabla
    for j in range(df[s].shape[0]):
        for c in range(len(names)):
            data = df[s].iloc[j,c]
            ws['{}{}'.format(column[c],j+2)] = data
wb.save(filename = output)
        